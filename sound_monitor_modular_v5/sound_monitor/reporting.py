from datetime import datetime

_OPENPYXL_AVAILABLE = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, numbers
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except Exception:
    _OPENPYXL_AVAILABLE = False

def has_openpyxl() -> bool:
    return _OPENPYXL_AVAILABLE

def compute_summary_stats(history: list) -> dict:
    n = len(history)
    total_time_s = 0.0; weighted_sum_L = 0.0
    peak_db = float("-inf"); peak_vol = float("-inf")
    max_dose = 0.0
    t_to_50 = None; t_to_100 = None
    if n >= 1:
        for i in range(n - 1):
            cur, nxt = history[i], history[i+1]
            dt = max(0.0, float(nxt["t_session"]) - float(cur["t_session"]))
            total_time_s += dt
            L = float(cur["L"]); weighted_sum_L += L * dt
            peak_db = max(peak_db, L)
            peak_vol = max(peak_vol, float(cur["vol_percent"]))
            max_dose = max(max_dose, float(cur["dose"]))
            if t_to_50 is None and float(cur["dose"]) >= 0.5: t_to_50 = float(cur["t_session"])
            if t_to_100 is None and float(cur["dose"]) >= 1.0: t_to_100 = float(cur["t_session"])
        last = history[-1]
        peak_db = max(peak_db, float(last["L"]))
        peak_vol = max(peak_vol, float(last["vol_percent"]))
        max_dose = max(max_dose, float(last["dose"]))
        if t_to_50 is None and float(last["dose"]) >= 0.5: t_to_50 = float(last["t_session"])
        if t_to_100 is None and float(last["dose"]) >= 1.0: t_to_100 = float(last["t_session"])
    avg_db = (weighted_sum_L / total_time_s) if total_time_s > 0 else 0.0
    return {
        "points": n,
        "total_time_s": total_time_s,
        "total_time_days": total_time_s / 86400.0,
        "avg_db": avg_db,
        "peak_db": peak_db if peak_db != float("-inf") else 0.0,
        "peak_vol": peak_vol if peak_vol != float("-inf") else 0.0,
        "max_dose": max_dose,
        "t_to_50_days": (t_to_50 / 86400.0) if t_to_50 is not None else 0.0,
        "t_to_100_days": (t_to_100 / 86400.0) if t_to_100 is not None else 0.0,
    }

def make_workbook(history: list, cfg: dict):
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl não disponível")
    wb = Workbook()
    ws = wb.active; ws.title = "Relatório"
    headers = ["timestamp_iso","t_sessao_s","modo","volume_%","nivel_dB","dose_0a1","zona","dose_diaria"]
    ws.append(headers)
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=c); cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")
    for row in history:
        ws.append([
            row["ts_iso"], float(row["t_session"]), row["mode"],
            int(round(float(row["vol_percent"]))),
            float(row["L"]), float(row["dose"]), row["zone"], float(row["daily"])
        ])
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 2).number_format = "0.0"
        ws.cell(r, 5).number_format = "0.00"
        ws.cell(r, 6).number_format = numbers.FORMAT_PERCENTAGE_00
        ws.cell(r, 8).number_format = numbers.FORMAT_PERCENTAGE_00
    widths = [20,14,12,12,12,12,16,16]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.auto_filter.ref = f"A1:H{ws.max_row}"; ws.freeze_panes = "A2"

    ws2 = wb.create_sheet(title="Resumo")
    from .reporting import compute_summary_stats as _css
    if not history:
        ws2["A1"] = "Sem dados na sessão."; ws2["A1"].font = Font(bold=True)
    else:
        summary = _css(history)
        ws2["A1"] = "Resumo da Sessão"; ws2["A1"].font = Font(bold=True)
        ws2["A2"] = f"Gerado em: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws2["A3"] = f"Perfil diário: {cfg['ref_db']:.0f} dB / 8h (3 dB)"
        labels = [
            ("Tempo total",          summary["total_time_days"], "[h]:mm:ss"),
            ("Média de dB",          summary["avg_db"],          "0.00"),
            ("Pico de dB",           summary["peak_db"],         "0.00"),
            ("Pico de volume (%)",   summary["peak_vol"],        "0"),
            ("Maior dose (sessão)",  summary["max_dose"],        numbers.FORMAT_PERCENTAGE_00),
            ("Tempo até 50% dose",   summary["t_to_50_days"],    "[h]:mm:ss"),
            ("Tempo até 100% dose",  summary["t_to_100_days"],   "[h]:mm:ss"),
        ]
        ws2["A5"] = "Métricas gerais"; ws2["A5"].font = Font(bold=True)
        row = 6
        for label, value, fmt in labels:
            ws2.cell(row=row, column=1, value=label)
            cval = ws2.cell(row=row, column=2, value=value); cval.number_format = fmt
            row += 1
        ws2.column_dimensions["A"].width = 26
        ws2.column_dimensions["B"].width = 18
    return wb
